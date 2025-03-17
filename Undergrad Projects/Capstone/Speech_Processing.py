# -*- coding: utf-8 -*-
"""
Created on Thu Aug 25 12:39:30 2022

@author: Indrajit
"""
from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

import subprocess
import datetime
import wolframalpha
import pyttsx3
import tkinter
import json
import random
import operator
import speech_recognition as sr
import datetime
import wikipedia
import webbrowser
import os
import winshell
import pyjokes
import feedparser
import smtplib
import ctypes
import time
import requests
import shutil
import sounddevice as sd
from twilio.rest import Client
from clint.textui import progress
from ecapture import ecapture as ec
from bs4 import BeautifulSoup
import win32com.client as wincl
from urllib.request import urlopen
import playsound as ps
import pygame
l_med=[]
l_day=[]
l_hr=[]
l_min=[]
engine = pyttsx3.init('sapi5')
voices = engine.getProperty('voices')
engine.setProperty('voice', voices[1].id)
def speak(audio):
    engine.say(audio)
    engine.runAndWait()
def med():
    file=pathlib.Path('D:\Capstone Resources\medicine_record.xlsx')
    if file.exists():
        pass
    else:
        file=Workbook()
        sheet=file.active
        sheet['A1']="Medicine Name"
        sheet['B1']="Hour"
        sheet['C1']="Minute"
        sheet['D1']='Day'
        file.save('D:\Capstone Resources\medicine_record.xlsx')
    def clear():
        medvalue.set('')
        hrvalue.set('')
        minvalue.set('')
        dayvalue.set('')
    def submit():
        med=medvalue.get()
        hr=hrvalue.get()
        mini=minvalue.get()
        day=dayvalue.get()
        l_med.append(med)
        l_hr.append(hr)
        l_min.append(mini)
        l_day.append(day)
        print(med)
        print(hr)
        print(mini)
        print(day)
        file=openpyxl.load_workbook('D:\Capstone Resources\medicine_record.xlsx')
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=med)
        sheet.cell(column=2,row=sheet.max_row,value=hr)
        sheet.cell(column=3,row=sheet.max_row,value=mini)
        sheet.cell(column=4,row=sheet.max_row,value=day)
        file.save(r'D:\Capstone Resources\medicine_record.xlsx')
        messagebox.showinfo('info','detail added!')
        medvalue.set('')
        hrvalue.set('')
        minvalue.set('')
        dayvalue.set('')

    root=Tk()
    root.title("Medicine Entry")
    root.geometry('700x400+300+200')
    root.resizable(False,False)
    root.configure(bg="#326273")
    Label(root,text="Please fill this form:",font="arial 13",bg="#326273",fg="#fff").place(x=20,y=20)
    Label(root,text="Medicine Name",font=23,bg="#326273",fg="#fff").place(x=50,y=100)
    Label(root,text="Hour",font=23,bg="#326273",fg="#fff").place(x=50,y=150)
    Label(root,text="Minute",font=23,bg="#326273",fg="#fff").place(x=50,y=200)
    Label(root,text="Day should take it",font=23,bg="#326273",fg="#fff").place(x=50,y=250)
    medvalue=StringVar()
    hrvalue=StringVar()
    minvalue=StringVar()
    dayvalue=StringVar()
    medentry=Entry(root,textvariable=medvalue,width=45,bd=2,font=20)
    hrentry=Entry(root,textvariable=hrvalue,width=45,bd=2,font=20)
    minentry=Entry(root,textvariable=minvalue,width=45,bd=2,font=20)
    dayentry=Entry(root,textvariable=dayvalue,width=45,bd=2,font=20)
    medentry.place(x=200,y=100)
    hrentry.place(x=200,y=150)
    minentry.place(x=200,y=200)
    dayentry.place(x=200,y=250)
    Button(root,text="Submit",bg="#326273",fg="white",width=15,height=2,command=submit).place(x=200,y=350)
    Button(root,text="Clear",bg="#326273",fg="white",width=15,height=2,command=clear).place(x=340,y=350)
    Button(root,text="Exit",bg="#326273",fg="white",width=15,height=2,command=lambda:root.destroy()).place(x=480,y=350)
    root.mainloop()
    
def wishMe():
    hour = int(datetime.datetime.now().hour)
    if hour>= 0 and hour<12:
        speak("Good Morning mam !")
  
    elif hour>= 12 and hour<18:
        speak("Good Afternoon mam !")  
  
    else:
        speak("Good Evening mam !") 
  
    assname =("Jarvis")
    speak("I am your Assistant")
    speak(assname)
     
 
def username():
    speak("What should i call you mam")
    uname = takeCommand()
    speak("Welcome Miss")
    speak(uname)
    columns = shutil.get_terminal_size().columns
     
    print("#####################".center(columns))
    print("Welcome Ms.", uname.center(columns))
    print("#####################".center(columns))
     
    speak("How can i Help you, Mam")
 
def takeCommand():
    
    r = sr.Recognizer()
     
    with sr.Microphone() as source:
        r.adjust_for_ambient_noise(source,duration=1)
        print("Listening...")
        r.pause_threshold = 1
        audio = r.listen(source)
  
    try:
        print("Recognizing...")   
        query = r.recognize_google(audio, language ='en-in')
        print(f"User said: {query}\n")
  
    except Exception as e:
        print(e)   
        print("Unable to Recognize your voice.") 
        return "None"
     
    return query
if __name__ == '__main__':
    clear = lambda: os.system('cls')
     
    # This Function will clean any
    # command before execution of this python file
    clear()
    wishMe()
    username()
     
    while True:
        cur_hr=datetime.datetime.now().hour
        cur_min=datetime.datetime.now().minute
        x_date = datetime.datetime.now()
        no = x_date.weekday()
        query = takeCommand().lower()
        if 'play music' in query or "play song" in query:
            speak("Here you go with music")
            # music_dir = "G:\\Song"
            #music_dir = "C://Users//Indrajit//Music//"
            #songs = os.listdir(music_dir)
            #print(songs)   
            #os.startfile(os.path.join(music_dir, songs[1]))
            #os.startfile("C://Users//Indrajit//Music//Rangisaari _ Jug Jugg Jeeyo _Team Naach Choreography _.mp3")
            #ps.playsound(os.path.join(music_dir, songs[1]))
            path = "D:\\Capstone Resources\\Music\\"
            all_mp3 = [os.path.join(path, f) for f in os.listdir(path) if f.endswith('.mp3')]
            randomfile = random.choice(all_mp3)
            pygame.mixer.init()
            pygame.mixer.music.load(randomfile)
            pygame.mixer.music.play()
        elif 'stop music' in query:
            pygame.mixer.music.stop()
        elif 'the time' in query:
            strTime = datetime.datetime.now().strftime("%m-%d-%Y %H:%M%p")
            speak(f"Maam, the time is {strTime}")
 
        elif 'how are you' in query:
            speak("I am fine, Thank you")
            speak("How are you, maam")
 
        elif 'fine' in query or "good" in query:
            speak("It's good to know that your fine")
 
        elif "change my name to" in query:
            query = query.replace("change my name to", "")
            assname = query
 
        elif "change name" in query:
            speak("What would you like to call me, maam")
            assname = takeCommand()
            speak("Thanks for naming me")
 
        elif "what's your name" in query or "What is your name" in query:
            speak("My friends call me")
            speak(assname)
            print("My friends call me", assname)
 
        elif 'exit' in query:
            speak("Thanks for giving me your time")
            exit()
 
        elif "who made you" in query or "who created you" in query:
            speak("I have been created by Akshata, Pubali, Aishwarya, Priyanka.")
             
        elif 'joke' in query:
            speak(pyjokes.get_joke())
 
 
        elif "who i am" in query:
            speak("If you talk then definitely your human.")
 
        elif "why you came to world" in query:
            speak("So that I can be your friend")
 
        elif "who are you" in query:
            speak("I am your virtual assistant")
        elif 'news' in query:
             
            # try:
            #     jsonObj = urlopen('''https://newsapi.org / v1 / articles?source = the-times-of-india&sortBy = top&apiKey =\\times of India Api key\\''')
            #     data = json.load(jsonObj)
            #     i = 1
                 
            #     speak('here are some top news from the times of india')
            #     print('''=============== TIMES OF INDIA ============'''+ '\n')
                 
            #     for item in data['articles']:
                     
            #         print(str(i) + '. ' + item['title'] + '\n')
            #         print(item['description'] + '\n')
            #         speak(str(i) + '. ' + item['title'] + '\n')
            #         i += 1
            # except Exception as e:
                 
            #     print(str(e))
            import requests
            from bs4 import BeautifulSoup
            url = 'https://www.bbc.com/news'
            response = requests.get(url)
            soup = BeautifulSoup(response.text, 'html.parser')
            headlines = soup.find('body').find_all('h3')
            unwanted = ['BBC World News TV', 'BBC World Service Radio',
            'News daily newsletter', 'Mobile app', 'Get in touch']
            for x in list(dict.fromkeys(headlines)):
                if x.text.strip() not in unwanted:
                    speak(x.text.strip())
                
 
        elif "don't listen" in query or "stop listening" in query:
            speak("for how much time you want to stop jarvis from listening commands")
            a = int(takeCommand())
            time.sleep(a)
            print(a)
 
        elif "where is" in query:
            query = query.replace("where is", "")
            location = query
            speak("User asked to Locate")
            speak(location)
            webbrowser.open("https://www.google.nl / maps / place/" + location + "")
 
        elif "camera" in query or "take a photo" in query:
            ec.capture(0,"Jarvis Camera", "img.jpg")
 
        elif "write a note" in query:
            speak("What should i write, sir")
            note = takeCommand()
            file = open('jarvis.txt', 'w')
            speak("Sir, Should i include date and time")
            snfm = takeCommand()
            if 'yes' in snfm or 'sure' in snfm:
                strTime = datetime.datetime.now().strftime("% H:% M:% S")
                file.write(strTime)
                file.write(" :- ")
                file.write(note)
            else:
                file.write(note)
         
        elif "show note" in query:
            speak("Showing Notes")
            file = open("jarvis.txt", "r")
            print(file.read())
            speak(file.read(6))
 
        elif "jarvis" in query:
             
            wishMe()
            speak("Jarvis 1 point o in your service Miss")
            speak(assname)
 
        elif "weather" in query:
            from bs4 import BeautifulSoup
            import requests
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}
            def weather(city):
                city = city.replace(" ", "+")
                res = requests.get(f'https://www.google.com/search?q={city}&oq={city}&aqs=chrome.0.35i39l2j0l4j46j69i60.6128j1j7&sourceid=chrome&ie=UTF-8', headers=headers)
                print("Searching...\n")
                soup = BeautifulSoup(res.text, 'html.parser')
                location = soup.select('#wob_loc')[0].getText().strip()
                time = soup.select('#wob_dts')[0].getText().strip()
                info = soup.select('#wob_dc')[0].getText().strip()
                weather = soup.select('#wob_tm')[0].getText().strip()
                speak(location)
                speak(time)
                speak(info)
                speak(weather+"°C")
            #speak("Enter the Name of City ->  ")
            #city = takeCommand()+" weather"
            city = "Bangalore"+" weather"
            weather(city)
            speak("Have a Nice Day:)") 
        elif "send message " in query:
                # You need to create an account on Twilio to use this service
                account_sid = 'Account Sid key'
                auth_token = 'Auth token'
                client = Client(account_sid, auth_token)
 
                message = client.messages \
                                .create(
                                    body = takeCommand(),
                                    from_='Sender No',
                                    to ='Receiver No'
                                )
 
                print(message.sid)
 
        elif "wikipedia" in query:
            webbrowser.open("wikipedia.com")
 
        elif "Good Morning" in query:
            speak("A warm" +query)
            speak("How are you Miss")
            speak(assname)
 
        elif "how are you" in query:
            speak("I'm fine, glad you me that")
        elif "what is" in query or "who is" in query:
             
            # Use the same API key
            # that we have generated earlier
            client = wolframalpha.Client("API_ID")
            res = client.query(query)
             
            try:
                print (next(res.results).text)
                speak (next(res.results).text)
            except StopIteration:
                print ("No results")
 
        # elif "medicine" in query and "take today" in query:
        #     if no==0:
        #         speak("You have to take 1 pill of Aducanumab at 10:00 am")
        #         speak("you have to take 2 pills of Galantamine at 6:00pm")
                
            
       # elif "medicine" in query and "now" in query:
        
            #now=datetime.datetime.now().time()
        elif "medicine entry" in query:
            med()
        import pandas as pd
        workbook = pd.read_excel('D:\Capstone Resources\medicine_record.xlsx')
        rows = len(workbook.axes[0])
        for i in range(0,rows):
            d=workbook['Day'].iloc[i]
            h=workbook['Hour'].iloc[i]
            m=workbook['Minute'].iloc[i]
            n=workbook['Medicine Name'].iloc[i]
            if d=='Monday':
                if (cur_hr==(int(h))) & (cur_min==(int(m))):
                    speak("you have to take")
                    speak(n)
            elif d=='Tuesday':
                if (cur_hr==(int(h))) & (cur_min==(int(m))):
                    speak("you have to take")
                    speak(n)
            elif d=='Wednesday':
                if (cur_hr==(int(h))) & (cur_min==(int(m))):
                    speak("you have to take")
                    speak(n)
            elif d=='Thursday':
                if (cur_hr==(int(h))) & (cur_min==(int(m))):
                    speak("you have to take")
                    speak(n)
            elif d=='Friday':
                if (cur_hr==(int(h))) & (cur_min==(int(m))):
                    speak("you have to take")
                    speak(n)
            elif d=='Saturday':
                if (cur_hr==(int(h))) & (cur_min==(int(m))):
                    speak("you have to take")
                    speak(n)
            elif d=='Sunday':
                if (cur_hr==(int(h))) & (cur_min==(int(m))):
                    speak("you have to take")
                    speak(n)
            
        # elif (cur_hr==10) & (cur_min==10) & (no==0):
        #     speak("it's Monday 10:10am") 
        #     speak("you have to take one pill of Aducanumab")
        # elif (cur_hr==10) &(cur_min==37) & (no==3):
        #     speak("It's Tuesday 6:21 pm")
        #     speak("You have to take one pill of Galantamine")
            # Command go here
            # For adding more commands